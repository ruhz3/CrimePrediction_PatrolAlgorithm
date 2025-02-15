
# ================================================
# artificial_case.py
# - CCTV-경찰서, 유동인구-거주인구로 선별한 그리드에 범죄 사례 생성
# ================================================


import numpy as np
import cv2
import random
import openpyxl
from const_data import *

''' ↓ 저장할 엑셀 파일 이름, 연도별 경향성 데이터 수 입력'''
FILE_NAME = "crime_data"
CRIME_TOTAL = [3000, 3000, 3000, 3000, 3000, 3000]  # 2014년 - 2019년

WIDTH = 59
HEIGHT = 32
CRIME_KEY = {0: 'Theft', 1: 'Assault', 2: 'Murder', 3: 'Rape', 4: 'Robber'}
YEAR_KEY = {0: '2014', 1: '2015', 2: '2016', 3: '2017', 4: '2018', 5: '2019'}

# 그리드 이미지를 배열로 변환
img = cv2.imread("img/grid.png", 0)
img = cv2.resize(dsize=(WIDTH, HEIGHT), src=img)
grids = []
result = np.zeros((HEIGHT, WIDTH, 3), dtype=np.uint8)
for x in range(WIDTH):
    for y in range(HEIGHT):
        if img[y][x] != 255:
            grids.append([y, x])

# 엑셀 파일 생성(연도별 시트 생성)
grd_switch = False
for year in range(6):
    wb = openpyxl.load_workbook(f'data/{FILE_NAME}_{YEAR_KEY[year]}.xlsx')
    print(f'{YEAR_KEY[year]}...')
    sheet = wb.active
    sheet['A1'] = "Type"
    sheet['B1'] = "Date-Time"
    sheet['C1'] = "Latitude"
    sheet['D1'] = "Longitude"
    sheet['E1'] = "Grid"
    case_num = CRIME_TOTAL[year]
    last_row = sheet.max_row

    # 통계자료에 맞춰 범죄사례를 생성
    for i in range(last_row + 1, last_row + 1 + case_num):
        # 범죄유형
        n = random.random()
        for c in range(5):
            if n < CRIME_STATS[year][c]:
                sheet.cell(row=i, column=1).value = f'{CRIME_KEY[c]}'
                crime = c
                break

        # 범죄날짜 및 시간
        month = random.randrange(1, 13)
        day = random.randrange(1, MONTH_DAYS[month] + 1)
        if month < 10:
            str_month = f'0{month}'
        else:
            str_month = f'{month}'
        if day < 10:
            str_day = f'0{day}'
        else:
            str_day = f'{day}'
        date = f'{YEAR_KEY[year]}-{str_month}-{str_day}'

        n = random.random()
        for k in range(8):
            if n < CRIME_TIME[crime][k]:
                hour = random.randrange(3*k, 3*(k+1))
                break
        if hour < 10:
            str_hour = f'0{hour}'
        else:
            str_hour = f'{hour}'
        time = str_hour + ":00:00"
        sheet.cell(row=i, column=2).value = date + " " + time

        # 범죄지점 : 그리드를 랜덤으로 선택하고, 그리드 내부에서 랜덤생성
        tt = int(hour/6)
        if grd_switch:
            gg = DANGER_GRID_AND
        else:
            gg = DANGER_GRID_OR
        grd_switch = (grd_switch == False)
        ll = len(gg[tt])
        ii = int(random.uniform(0, ll))
        idx = gg[tt][ii]
        x, y = grids[idx]
        gx, gy = O_LATITUDE - x * LAT_GAP, O_LONGTITUDE + y * LONG_GAP
        sheet.cell(row=i, column=3).value = random.uniform(gx - LAT_GAP, gx)
        sheet.cell(row=i, column=4).value = random.uniform(gy, gy + LONG_GAP)
        sheet.cell(row=i, column=5).value = idx

    wb.save(f'data\{FILE_NAME}_{YEAR_KEY[year]}.xlsx')
    wb.close()

# 파일 저장
print('Artificial data added!')



